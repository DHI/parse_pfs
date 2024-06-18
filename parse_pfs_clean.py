import mikeio
import pandas as pd
import os
from pyproj import CRS
from mikeio.pfs._pfssection import PfsSection
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font

class utils:
    def __init__(self):
        self.units = {"Eddy": "m²/s",
                 "Density": "kg/m³",
                 "Depth": "m",
                 "Smagorinsky coefficient": "",
                 "Drag": "",
                 "Roughness height": "m",
                 "Sediment depth": "mm",
                 "Ice roughness height": "m",
                 "Precipitation": "mm/d",
                 "Evaporation": "mm/d",
                 "Scaled eddy viscosity": "",
                 "Dispersion coefficient": "m²/s",
                 "Light extinction": "1/m",
                 "Short wave radiation": "W/m²",
                 "Long wave radiation": "W/m²",
                 "Air temperature": "°C",
                 "Relative humidity": "%",
                 "Clearness coefficient": "",
                 "Thermal conductivity": "W/m/°C",
                 "Ground temperature": "°C",
                 "Temperature": "°C",
                 "Turbulent kinetic energy": "m²/s²",
                 "Specific dissipation rate": "1/s",
                 "Dissipation of turbulent kinetic energy": "m²/s³",
                 "Grain diameter": "mm",
                 "Chezy number": "m^(1/2)/s",
                 "Manning number": "m^(1/3)/s",
                 "Alluvial resistance": "m^(1/3)/s",
                 "Concentration": "g/m³",
                 "Thickness": "m",
                 "Composition": "%",
                 "Speedup factor": "",
                 "Angle of repose": "°",
                 "Added rate of bed level change": "m/d",
                 "Maximum allowed bed level": "m",
                 "Distribution factor": "",
                 }
    
    def _parse_format(self, data, variable, key="constant_value", has_soft_time=False):
        output = {}
        
        if data["format"] == 0:
            output[f"{variable} format"] = "Constant"
            output[f"{variable} value"] = f"{data[key]} {self.units[variable]}"
        if data["format"] == 2:
            output[f"{variable} format"] = "Constant in time, varying in space"
        elif data["format"] == 3:
            output[f"{variable} format"] = "Variying in time and space"
        if has_soft_time:
            output["Soft time interval"] = f"{data['soft_time_interval']} s"
        return output

    def _parse_coordinates(self, projection):
        try:
            crs = CRS.from_string(projection)
            coordinate_system = crs.name
        except:
            coordinate_system = projection
        return coordinate_system

class PFS_Domain(utils):
    def __init__(self, domain):
        super().__init__()
        self.domain = domain
        self.mesh_file = domain["file_name"]
        self.coordinate_system = self._parse_coordinates(domain["coordinate_type"])
        self.minimum_depth = domain["minimum_depth"]
        self.datum_shift = domain["datum_depth"]
        self.vertical_mesh = self.parse_vertical_mesh()
        self.boundaries = self.parse_boundaries(domain["BOUNDARY_NAMES"])
        self.attributes = {"Mesh file": self.mesh_file,
                           "Coordinate system": self.coordinate_system,
                           "Minimum depth": str(self.minimum_depth) + " m",
                           "Datum shift": str(self.datum_shift) + " m",
                           "Vertical mesh": self.vertical_mesh
                           }
    
    def parse_vertical_mesh(self):
        mesh_type = self.domain["vertical_mesh_type_overall"]
        n_sigma_layers = self.domain["number_of_layers"]
        sigma_depth = self.domain["z_sigma"]
        sigma_type = self.domain["vertical_mesh_type"]
        sigma_layer_thickness = self.domain["layer_thickness"]
        sigma_c = self.domain["sigma_c"]
        sigma_theta = self.domain["theta"]
        sigma_b = self.domain["b"]
        n_z_layers = self.domain["number_of_layers_zlevel"]
        z_type = self.domain["vertical_mesh_type_zlevel"]
        z_constant = self.domain["constant_layer_thickness_zlevel"]
        z_variable = self.domain["variable_layer_thickness_zlevel"]
        z_level_bathymetry_adjustment = self.domain["type_of_bathymetry_adjustment"]
        z_minimum_layer_thickness = self.domain["minimum_layer_thickness_zlevel"]
        if mesh_type == 1:
            vertical_mesh = {"Vertical mesh type": self._parse_sigma(sigma_type, n_sigma_layers, sigma_c, sigma_theta, sigma_b, sigma_depth)}
        elif mesh_type == 2:
            sigma = self._parse_sigma(sigma_type, n_sigma_layers, sigma_c, sigma_theta, sigma_b, sigma_depth, has_z=True)
            z = self._parse_zlevel(z_type, n_z_layers, z_constant, z_level_bathymetry_adjustment, z_minimum_layer_thickness)
            vertical_mesh = {"Vertical mesh type": "{sigma} at the top and {z} at the bottom".format(sigma=sigma, z=z[0])} | z[1]
        return vertical_mesh

    def parse_boundaries(self, boundaries):
        output = {"CODE_1": {"Name": "Land Boundary"}}
        for key in boundaries.keys():
            if isinstance(boundaries[key], PfsSection):
                output = output | {key: {"Name": boundaries[key]["name"]}}
        return output
    
    def _parse_sigma(self, sigma, n_layers, c, theta, b, sigma_depth, has_z=False):
        if sigma == 1:
            output = f"σ-coordinate with {n_layers} equidistant layers"
        elif sigma == 2:
            output = f"σ-coordinate with {n_layers} layers with specified thicknesses"
        elif sigma == 3:
            output = f"σ-coordinate with {n_layers} layers with specified distribution parameters (σ_c = {c}, θ = {theta}, b = {b})"
        if has_z:
            output = output + f" (Sigma depth = {sigma_depth} m)"
        return output

    def _parse_zlevel(self, z_type, n_layers, constant, bathymetry_adjustment, min_thickness):
        if z_type == 1:
            output = [f"z-coordinate with {n_layers} equidistant layers (constant thickness = {constant})"]
        elif z_type == 2:
            output = [f"z-coordinate with {n_layers} layers with specified thicknesses"]
        if bathymetry_adjustment == 1:
            output.append({"Type of bathymetry adjustment": "Simple adjustment"})
        elif bathymetry_adjustment == 2:
            output.append({"Type of bathymetry adjustment": "Advanced adjustment",
                           "Minimum layer thickness": str(min_thickness) + " m"})
        return output

class PFS_Time:
    def __init__(self, time):
        self.time = time
        self.start_time = datetime(*time["start_time"])
        self.time_step_interval = f'{time["time_step_interval"]} s'
        self.number_of_time_steps = time["number_of_time_steps"]
        self.end_time = self.start_time + timedelta(seconds=time["number_of_time_steps"]*time["time_step_interval"])
        self.attributes = {"Start Time": self.start_time.strftime("%Y-%m-%d %I:%M:%S %p"),
                            "Time Step Interval": self.time_step_interval,
                            "Number of Time Steps": str(self.number_of_time_steps),
                            "End Time": self.end_time.strftime("%Y-%m-%d %I:%M:%S %p")
                          }

class PFS_Equation:
    def __init__(self, equation):
        self.equation = equation
        self.formulation = equation["formulation"]
        if self.formulation not in [2, 4]:
            raise ValueError("Unknown Governing Equation")
        self.attributes = "Navier-Stokes equations" if self.formulation == 2 else "Shallow water equations" if self.formulation == 4 else None

class PFS_Solution_Technique:
    def __init__(self, solution_technique, module):
        self.solution_technique = solution_technique
        self.module = module
        integrations = {1: "Low order, fast algorithm", 2: "Higher order"}
        Reimann_solvers = {0: None, 3: "HLLC"}
        Filtering_options = {0: None, 1: "Explicit filter"}
        self.time_integration = integrations[solution_technique["scheme_of_time_integration"]] 
        self.space_discretization = integrations[solution_technique["scheme_of_space_discretization_horizontal"]]
        if module == "hydrodynamic":
            self.type_of_filtering = Filtering_options[solution_technique["type_of_filtering"]]
            self.filtering_coefficient = solution_technique["filtering_coefficient"]
            self.type_of_Riemann_solver = Reimann_solvers[solution_technique["type_of_Riemann_solver"]]
            self.Riemann_factor = solution_technique["Riemann_factor"]
            self.CFL_critical_HD = solution_technique["CFL_critical_HD"]
            self.CFL_critical_AD = solution_technique["CFL_critical_AD"]
            self.dt_min_HD = solution_technique["dt_min_HD"]
            self.dt_max_HD = solution_technique["dt_max_HD"]
            self.dt_min_AD = solution_technique["dt_min_AD"]
            self.dt_max_AD = solution_technique["dt_max_AD"]
            self.attributes = {"Hydrodynamic solver": {"Time integration scheme": self.time_integration,
                                                      "Space integration scheme": self.space_discretization,
                                                      "Maximum time step": self.dt_max_HD,
                                                      "Minimum time step": self.dt_min_HD,
                                                      "Critical CFL": self.CFL_critical_HD}}
            if self.type_of_filtering is not None:
                self.attributes["Hydrodynamic solver"] = self.attributes["Hydrodynamic solver"] | {"Type of filter": self.type_of_filtering,
                                                    "Filter coefficient": self.filtering_coefficient}
            if self.type_of_Riemann_solver is not None:
                self.attributes["Hydrodynamic solver"] = self.attributes["Hydrodynamic solver"] | {"Reimann solver": self.type_of_Riemann_solver,
                                                    "Reimann factor": self.Riemann_factor}

            self.attributes = self.attributes | {"Transport solver": {"Maximum time step": self.dt_max_AD,
                                                                     "Minimum time step": self.dt_min_AD,
                                                                     "Critical CFL": self.CFL_critical_AD}}
            
        elif module == "salinity_temperature":
            self.attributes = {"Time integration scheme": self.time_integration,
                               "Space integration scheme": self.space_discretization}
            
class PFS_Eddy(utils):
    def __init__(self, eddy):
        super().__init__()
        self.eddy = eddy
        functions = {0: self._parse_no_eddy, 1: self._parse_constant, 3: self._parse_smagorinsky, 4: self._parse_log_law, 5: self._parse_two_equation}
        keys = {0: "CONSTANT_EDDY_FORMULATION", 1: "CONSTANT_EDDY_FORMULATION", 3: "SMAGORINSKY_FORMULATION", 4: "LOG_LAW_FORMULATION", 5: "K_EPSILON_FORMULATION"}
        horizontal = eddy["HORIZONTAL_EDDY_VISCOSITY"]
        vertical = eddy["VERTICAL_EDDY_VISCOSITY"]
        self.horizontal = functions[horizontal["type"]](horizontal[keys[horizontal["type"]]])
        self.vertical = functions[vertical["type"]](vertical[keys[vertical["type"]]])
        self.attributes = {"Horizontal Eddy Viscosity": self.horizontal, "Vertical Eddy Viscosity": self.vertical}


    def _parse_no_eddy(self, eddy):
        return "No eddy"

    def _parse_constant(self, eddy):
        output = {"Eddy type": "Constant"}
        output.update(self._parse_format(eddy, variable="Eddy viscosity", key="constant_value"))
        try:
            output.update(self.__parse_damping(eddy))
        except:
            pass
        return output
    
    def _parse_smagorinsky(self, eddy):
        output = {"Eddy type": "Smagorinsky formulation"}
        output.update(self._parse_format(eddy, variable="Smagorinsky coefficient", key="constant_value"))
        output.update(self.__parse_min_max(eddy))
        return output
    
    def _parse_log_law(self, eddy):
        output = {"Eddy type": "Log law formulation"}
        output.update(self.__parse_min_max(eddy))
        output.update(self.__parse_damping(eddy))
        return output
    
    def _parse_two_equation(self, eddy):
        output = {"Eddy type": "Two-equation turbulence model"}
        output.update(self.__parse_min_max(eddy))
        return output
    
    def __parse_damping(self, eddy):
        if eddy["Ri_damping"] == 0:
            return {"Damping": "Not included"}
        elif eddy["Ri_damping"] == 1:
            return {"Damping constant a": eddy["Ri_a"], "Damping constant b": eddy["Ri_b"]}
        
    def __parse_min_max(self, eddy):
        return {"Minimum eddy viscosity": eddy["minimum_eddy_viscosity"], "Maximum eddy viscosity": eddy["maximum_eddy_viscosity"]}         

class Bed_Resistence(utils):
    def __init__(self, bed):
        super().__init__()
        self.bed = bed
        beds = {0: "No bed resistance", 2: "Quadratic drag coefficient", 5: "Roughness height", 7: "Wave induced bed resistance"}
        keys = {0: None, 2: "DRAG_COEFFICIENT", 5: "ROUGHNESS", 7: "WAVE_INDUCED_ROUGHNESS"}
        vars = {2: "Drag coefficient", 5: "Roughness height", 7: "Sediment depth"}
        self.type = beds[bed["type"]]
        self.attributes = {"Bed resistance type": self.type}
        if bed["type"] > 0:
            self.attributes.update(self._parse_format(bed[keys[bed["type"]]], variable=vars[bed["type"]], key="constant_value"))
        if bed["type"] == 7:
            self.attributes.update(self._parse_wave_correction(bed["WAVE_INDUCED_ROUGHNESS"]))
            
    def _parse_wave_correction(self, wave):
        if wave["type_of_wave_height_correction"] == 0:
            return {"Wave correction": "Not included"}
        else:
            return {"Wave correction type": "Wave height correction",
                    "Wave height over depth limit": wave["wave_height_over_depth_limit"],
                    "Minimum water depth for including waves": wave["minimum_waterdepth_for_including_waves"]}

class Wind_Forcing(utils):
    def __init__(self, wind):
        super().__init__()
        self.wind = wind
        if wind["type"] == 0:
            self.attributes = {"Wind forcing": "Not included"}
        else:
            forcing = self._parse_forcing(wind)
            friction = self._parse_friction(wind["WIND_FRICTION"])
            self.attributes = {"Wind forcing": forcing, "Wind friction": friction}
        
    def _parse_forcing(self, wind):
        if wind["format"] == 0:
            return {"Wind forcing format": "Constant",
                    "Constant wind speed": str(wind["constant_speed"]) + " m/s",
                    "Constant wind direction": str(wind["constant_direction"]) + " degrees",
                    "Soft time interval": str(wind["soft_time_interval"]) + " s"
                    }
        elif wind["format"] == 1:
            return {"Wind forcing format": "Varying in time, constant in space",
                    "Soft time interval": str(wind["soft_time_interval"]) + " s"
                    }
        elif wind["format"] == 3:
            return {"Wind forcing format": "Varying in time and space",
                    "Neutral pressure": str(wind["neutral_pressure"]) + " hPa",
                    "Soft time interval": str(wind["soft_time_interval"]) + " s"
                    }
        
    def _parse_friction(self, friction):
        if friction["type"] == 0:
            return {"Wind friction type": "Constant",
                    "Constant wind friction": friction["constant_friction"]
                    }
        elif friction["type"] == 1:
            return {"Wind friction type": "Function of wind speed",
                    "Lower limit for linear variation": f"{friction['linear_friction_low']} at {friction['linear_speed_low']} m/s",
                    "Upper limit for linear variation": f"{friction['linear_friction_high']} at {friction['linear_speed_high']} m/s"
                    }
        
class PFS_Source(utils):
    def __init__(self, sources_value):
        super().__init__()
        self.sources_value = sources_value
        n_sources = sources_value["number_of_sources"]
        self._parse_Sources(n_sources)
        self.attributes = {"Number of sources": self.n_sources, **self.sources_info}

    def _parse_Sources(self, n_sources):
        self.sources_info = {}
        self.n_sources = 0
        for i in range(n_sources):
            source = self.sources_value[f"SOURCE_{i+1}"]
            if source["include"] == 0:
                continue
            self.sources_info[f"Source {i+1}"] = {"Name": source["Name"],
                                                  "Location": self._parse_location(source),
                                                  "Hydrodynamic Information": self._parse_type(source)}
            self.n_sources += 1
        return self.sources_info
        
    def _parse_location(self, source):
        location = {"Map projection": self._parse_coordinates(source["coordinate_type"])}
        if source["interpolation_type"] == 0:
            location.update({"Coordinates": [source["coordinates"][0],
                                             source["coordinates"][1],
                                             "Vertical layer number {layer}".format(layer=source["layer"])]})
        elif source["interpolation_type"] == 1:
            location.update({"Coordinates": [source["coordinates"][0],
                                             source["coordinates"][1],
                                             "{Z} m below surface".format(Z=source["coordinates"][2])]})
        elif source["interpolation_type"] == 2:
            location.update({"Coordinates": [source["coordinates"][0],
                                             source["coordinates"][1],
                                             "{Z} m above bed".format(Z=source["coordinates"][2])]})
        elif source["interpolation_type"] == 3:
            location.update({"Coordinates": [source["coordinates"][0],
                                             source["coordinates"][1],
                                             "Z = {Z} m".format(Z=source["coordinates"][2])]})
        return location
    
    def _parse_type(self, source):
        if source["type"] == 1:
            source_info = {"Source type": "Simple source"}
            if source["format"] == 0:
                source_info["Source format"] = "Constant discharge"
                source_info["Discharge"] = str(source["constant_value"]) + " m³/s"
            elif source["format"] == 1:
                source_info["Source format"] = "Varying in time"
            elif source["format"] == 4:
                source_info["Source format"] = "Rating curve"
        elif source["type"] == 2:
            source_info = {"Source type": "Standard source"}
            if source["format"] == 0:
                source_info["Source format"] = "Constant discharge"
                source_info["Discharge"] = str(source["constant_values"][0]) + " m³/s"
                source_info["u-Velocity"] = str(source["constant_values"][1]) + " m/s"
                source_info["v-Velocity"] = str(source["constant_values"][2]) + " m/s"
                source_info["w-Velocity"] = str(source["constant_values"][3]) + " m/s"
            elif source["format"] == 1:
                source_info["Source format"] = "Varying in time"
        elif source["type"] == 3:
            source_info = {"Source type": "Connected source"}
            if source["format"] == 0:
                source_info["Source format"] = "Constant discharge"
                source_info["u-Velocity"] = str(source["constant_values"][1]) + " m/s"
                source_info["v-Velocity"] = str(source["constant_values"][2]) + " m/s"
            elif source["format"] == 1:
                source_info["Source format"] = "Varying in time"
        elif source["type"] == 4:
            source_info = {"Source type": "Jet"}
            if source["format"] == 0:
                source_info["Source format"] = "Constant discharge"
                source_info["Discharge"] = str(source["constant_value"]) + " m³/s"
            elif source["format"] == 1:
                source_info["Source format"] = "Varying in time"
            source_info["Jet diameter"] = str(source["diameter"]) + " m"
            source_info["Jet horizontal direction angle"] = str(source["sigma"]) + "°"
            source_info["Jet vertical direction angle"] = str(source["theta"]) + "°"
            source_info["Jet maximum travel distance"] = str(source["maximum_distance"]) + " m"
            if source["upstream"] == 1:
                source_info["Jet minimum upstream distance"] = str(source["distance_upstream"]) + " m"
        return source_info

class PFS_Boundaries(utils):
    def __init__(self, boundary_conditions, boundaries):
        super().__init__()
        self.boundaries = boundaries
        self.boundary_conditions = boundary_conditions
        self._parse_Boundary_Conditions()

    def _parse_Boundary_Conditions(self):
        for key in self.boundaries.keys():
            boundary = self.boundary_conditions[key]
            tmp = {}
            if boundary["type"] == 1:
                tmp["Type"] = "Land (zero normal velocity)"
                tmp["Wall friction"] = "Not included" if boundary["type_resistance"] == 0 else str(boundary["resistance_coefficient"]) + " m"
            elif boundary["type"] == 2:
                tmp["Type"] = "Land (zero velocity)"
            elif boundary["type"] == 4:
                tmp["Type"] = "Specified velocity"
                tmp = self.__parse_boundary_types(boundary, tmp, "velocity")
            elif boundary["type"] == 5:
                tmp["Type"] = "Specified flux"
                tmp = self.__parse_boundary_types(boundary, tmp, "flux")
            elif boundary["type"] == 6:
                tmp["Type"] = "Specified water level"
                self.__parse_boundary_types(boundary, tmp, "level")
            elif boundary["type"] == 7:
                tmp["Type"] = "Specified discharge"
                tmp = self.__parse_boundary_types(boundary, tmp, "discharge")
            elif boundary["type"] == 9:
                tmp["Type"] = "Free outflow"
            elif boundary["type"] == 12:
                tmp["Type"] = "Flather boundary"
                #TODO: Add Flather boundary conditions
            self.boundaries[key]["Hydrodynamic Information"] = tmp

    def __parse_boundary_types(self, boundary, tmp, var):
        if var == "velocity":
            vars = ["u-Velocity", "v-Velocity"]
            unit = "m/s"
        elif var == "flux":
            vars = ["p-flux", "q-flux"]
            unit = "m³/s/m"
        if var in ["velocity", "flux"] and boundary["format"] == 0:
            tmp["Format"] = "Constant"
            tmp[vars[0]] = str(boundary["constant_values"][0]) + f" {unit}"
            tmp[vars[1]] = str(boundary["constant_values"][1]) + f" {unit}"
            tmp["Type of vertical profile"] = "Uniform profile" if boundary["type_of_vertical_profile"] == 1 else "Logarithmic profile"
        elif var in ["level"] and boundary["format"] == 0:
            tmp["Format"] = "Constant"
            tmp["Water Level"] = str(boundary["constant_value"]) + " m"
        elif var in ["velocity", "flux"] and boundary["format"] == 1:
            tmp["Format"] = "Varying in time, constant along boundary"
            tmp["Type of vertical profile"] = "Uniform profile" if boundary["type_of_vertical_profile"] == 1 else "Logarithmic profile"
            tmp["Time interpolation type"] = "Linear" if boundary["type_of_time_interpolation"] == 1 else "Piecewise cubic"
        elif var in ["level"] and boundary["format"] == 1:
            tmp["Format"] = "Varying in time, constant along boundary"
            tmp["Time interpolation type"] = "Linear" if boundary["type_of_time_interpolation"] == 1 else "Piecewise cubic"
        elif var in ["velocity", "flux"] and boundary["format"] == 2:
            tmp["Format"] = "Varying in time and along boundary"
            tmp["Time interpolation type"] = "Linear" if boundary["type_of_time_interpolation"] == 1 else "Piecewise cubic"
            tmp["Space interpolation type"] = "Normal" if boundary["type_of_space_interpolation"] == 1 else "Reverse order"
        elif var in ["level"] and boundary["format"] == 2:
            tmp["Format"] = "Varying in time and along boundary"
            tmp["Time interpolation type"] = "Linear" if boundary["type_of_time_interpolation"] == 1 else "Piecewise cubic"
            tmp["Space interpolation type"] = "Normal" if boundary["type_of_space_interpolation"] == 1 else "Reverse order"
        elif var in ["level"] and boundary["format"] == 4:
            tmp["Format"] = "Rating curve"
        elif var in ["discharge"] and boundary["format"] == 0:
            tmp["Format"] = "Constant"
            tmp["Discharge"] = str(boundary["constant_value"]) + " m³/s"
            tmp["Approach"] = "Weak formulation" if boundary["approach"] == 1 else "Strong formulation"
        elif var in ["discharge"] and boundary["format"] == 1:
            tmp["Format"] = "Varying in time, constant along boundary"
            tmp["Approach"] = "Weak formulation" if boundary["approach"] == 1 else "Strong formulation"
            tmp["Time interpolation type"] = "Linear" if boundary["type_of_time_interpolation"] == 1 else "Piecewise cubic"
            tmp["Type of vertical profile"] = "Uniform profile" if boundary["type_of_vertical_profile"] == 1 else "Logarithmic profile"
        elif var in ["discharge"] and boundary["format"] == 4:
            tmp["Format"] = "Rating curve"
            tmp["Approach"] = "Weak formulation" if boundary["approach"] == 1 else "Strong formulation"
            tmp["Type of vertical profile"] = "Uniform profile" if boundary["type_of_vertical_profile"] == 1 else "Logarithmic profile"
            
        tmp["Type of soft start"] = "Linear variation" if boundary["type_of_soft_start"] == 1 else "Sine variation"
        tmp["Soft time interval"] = str(boundary["soft_time_interval"]) + " s"
        if var in ["velocity", "flux"]:
            tmp["Reference {var}".format(var=vars[0])] = str(boundary["reference_values"][0]) + f" {unit}"
            tmp["Reference {var}".format(var=vars[1])] = str(boundary["reference_values"][1]) + f" {unit}"
        elif var in ["level"]:
            tmp["Reference water level"] = str(boundary["reference_value"]) + " m"
            tmp["Coriolis correction for boundary data"] = "Included" if boundary["type_of_coriolis_correction"] == 1 else "Not included"
            tmp["Wind correction for boundary data"] = "Included" if boundary["type_of_wind_correction"] == 1 else "Not included"
            tmp["Pressure correction for boundary data"] = "Included" if boundary["type_of_pressure_correction"] == 1 else "Not included"
            tmp["Radiation stress correction for boundary data"] = "Included" if boundary["type_of_radiation_stress_correction"] == 1 else "Not included"
        return tmp

class PFS_Salinity_Temperature(utils):
    def __init__(self, salinity_temperature, boundaries, sources):
        super().__init__()
        self.salinity_temperature = salinity_temperature
        self.boundaries = boundaries
        self.sources = sources
        self.equation = self.salinity_temperature["EQUATION"]
        self.solution_technique = self.salinity_temperature["SOLUTION_TECHNIQUE"]
        self.dispersion = self.salinity_temperature["DIFFUSION"]
        self.heat_exchange = self.salinity_temperature["HEAT_EXCHANGE"]
        self.precipitation_evaporation = self.salinity_temperature["PRECIPITATION_EVAPORATION"]
        self.infiltration = self.salinity_temperature["INFILTRATION"]
        self.sources_conditions = self.salinity_temperature["SOURCES"]
        self.initial_conditions = self.salinity_temperature["INITIAL_CONDITIONS"]
        self.boundary_conditions = self.salinity_temperature["BOUNDARY_CONDITIONS"]
        self.attrubutes = {**self._parse_equation(),
                           **PFS_Solution_Technique(self.solution_technique, module="salinity_temperature").attributes,
                            **self._parse_Temperature_Salinity(self.salinity_temperature)
                           }
        self._parse_sources()
        self._parse_boundaries()
        
    def _parse_sources(self):
        for i in range(self.sources["Number of sources"]):
            source = self.sources_conditions[f"SOURCE_{i+1}"]
            types = {1: "Specified temperature", 2: "Excess temperature (Addition)", 3: "Excess temperature (Multiplication)"}
            self.sources[f"Source {i+1}"]["Temperature Information"] = {"Type": types[source["type_of_temperature"]],
                                                                        **self._parse_format(source["TEMPERATURE"], variable="Temperature")}
        
    def _parse_boundaries(self):
        for key in self.boundaries.keys():

            temperature = self.boundary_conditions[key]["TEMPERATURE"]
            if len(temperature.keys()) == 0:
                continue
            if temperature["type"] == 1:
               output = {"Type": "Land"}
            elif temperature["type"] == 2:
                output = {"Type": "Specified temperature",
                            **self._parse_format(temperature, variable="Temperature"),
                            "Type of soft start": "Linear variation" if temperature["type_of_soft_start"] == 1 else "Sine variation"}
                if temperature["format"] == 1:
                    output["Type of interpolation in time"] = "Linear" if temperature["type_of_time_interpolation"] == 1 else "Piecewise cubic"
                if temperature["format"] == 3:
                    output["Type of interpolation in time"] = "Linear" if temperature["type_of_time_interpolation"] == 1 else "Piecewise cubic"
                    output["Type of interpolation in space"] = "Normal" if temperature["type_of_space_interpolation"] == 1 else "Reverse order"
            elif temperature["type"] == 3:
                output = {"Type": "Zero gradient"}
            self.boundaries[key]["Temperature Information"] = output

    def _parse_equation(self):
        return{"Minimum temperature": str(self.equation["minimum_temperature"]) + " °C",
              "Maximum temperature": str(self.equation["maximum_temperature"]) + " °C",
              "Minimum salinity": str(self.equation["minimum_salinity"]) + " PSU",
              "Maximum salinity": str(self.equation["maximum_salinity"]) + " PSU"}

    def _parse_Temperature_Salinity(self, temperature_salinity):
        output = {}
        output["Dispersion"] = self.__parse_diffusion()
        output["Heat Exchange"] = self.__parse_heat_exchange()
        output["Precipitation - Evaporation"] = self.__parse_precipitation_evaporation()
        output["Infiltration"] = self.__parse_infiltration()
        output["Initial Conditions"] = self.__parse_initial_conditions()
        return output
    
    def __parse_diffusion(self):
        horizontal = self.dispersion["HORIZONTAL_DIFFUSION"]
        vertical = self.dispersion["VERTICAL_DIFFUSION"]
        types = {0: "No dispersion", 1: "SCALED_EDDY_VISCOSITY", 2: "DIFFUSION_COEFFICIENT"}
        variables = {1: "Scaled eddy viscosity", 2: "Dispersion coefficient"}
        keys = {1: "sigma", 2: "constant_value"}
        horizontal_diffusion = "Not included" if horizontal["type"] == 0 else {variables[horizontal["type"]]: self._parse_format(horizontal[types[horizontal["type"]]], variable=variables[horizontal["type"]], key=keys[horizontal["type"]])}
        vertical_diffusion = "Not included" if vertical["type"] == 0 else {variables[vertical["type"]]: self._parse_format(vertical[types[vertical["type"]]], variable=variables[vertical["type"]], key=keys[vertical["type"]])}
        return {"Horizontal Dispersion": horizontal_diffusion, "Vertical Dispersion": vertical_diffusion}

    def __parse_heat_exchange(self):
        if self.heat_exchange["type"] == 0:
            output = {"Heat exchange": "Not included"}
        elif self.heat_exchange["type"] == 1:
            # Latent heat
            latent = {"Conatant in Dalton's law": self.heat_exchange["Daltons_law_A"],
                   "Wind coefficient in Dalton's law": self.heat_exchange["Daltons_law_B"],
                   "Critical wind speed": str(self.heat_exchange["latent_heat_critical_wind_speed"]) + " m/s"}
            # Sensible heat
            sensible = {"Tranfer coefficient for heating": self.heat_exchange["sensible_heat_transfer_coefficient_heating"],
                        "Transfer coefficient for cooling": self.heat_exchange["sensible_heat_transfer_coefficient_cooling"],
                        "Critical wind speed": str(self.heat_exchange["sensible_heat_critical_wind_speed"]) + " m/s"}
            # Short wave radiation
            if self.heat_exchange["type_of_short_wave_radiation"] == 1:
                short_wave = {"Formulation": "Empirical",
                              "Sun constant, a in Angstroms's law": self.heat_exchange["Angstroms_law_A"],
                              "Sun constant, b in Angstroms's law": self.heat_exchange["Angstroms_law_B"],
                              "Displacement (summer time)": str(self.heat_exchange["displacement_hours"]) + " h",
                              "Standard meridian for time zone": self.heat_exchange["standard_meridian"],
                              "Beta in Beer's law": self.heat_exchange["Beers_law_beta"],
                              "Type of abosrption in water column": "Normalized" if self.heat_exchange["type_of_solar_radiation"] == 1 else "Not normalized"}
            elif self.heat_exchange["type_of_short_wave_radiation"] == 2:
                short_wave = {"Formulation": "Specified solar radiation",
                              "Displacement (summer time)": str(self.heat_exchange["displacement_hours"]) + " h",
                              "Standard meridian for time zone": self.heat_exchange["standard_meridian"],
                              "Beta in Beer's law": self.heat_exchange["Beers_law_beta"],
                              "Type of abosrption in water column": "Normalized" if self.heat_exchange["type_of_solar_radiation"] == 1 else "Not normalized"}
                short_wave = short_wave | {"Light extinction coefficient": self._parse_format(self.heat_exchange["LIGHT_EXTINCTION"], "Light extinction")}
                short_wave = short_wave | {"Radiation data": self._parse_format(self.heat_exchange["SHORT_WAVE_RADIATION_DATA"], "Short wave radiation")}
            elif self.heat_exchange["type_of_short_wave_radiation"] == 3:
                short_wave = {"Formulation": "Specified net short wave radiation",
                              "Beta in Beer's law": self.heat_exchange["Beers_law_beta"],
                              "Type of abosrption in water column": "Normalized" if self.heat_exchange["type_of_solar_radiation"] == 1 else "Not normalized"}
                short_wave = short_wave | {"Light extinction coefficient": self._parse_format(self.heat_exchange["LIGHT_EXTINCTION"], "Light extinction")}
                short_wave = short_wave | {"Radiation data": self._parse_format(self.heat_exchange["SHORT_WAVE_RADIATION_DATA"], "Short wave radiation")}
            # Long wave radiation
            if self.heat_exchange["type_of_long_wave_radiation"] == 1:
                long_wave = {"Formulation": "Empirical"}
            elif self.heat_exchange["type_of_long_wave_radiation"] == 2:
                long_wave = {"Formulation": "Specified atmospheric radiation"}
                long_wave = long_wave | {"Radiation data": self._parse_format(self.heat_exchange["LONG_WAVE_RADIATION_DATA"], "Long wave radiation")}
            elif self.heat_exchange["type_of_long_wave_radiation"] == 3:
                long_wave = {"Formulation": "Specified net long wave radiation"}
                long_wave = long_wave | {"Radiation data": self._parse_format(self.heat_exchange["LONG_WAVE_RADIATION_DATA"], "Long wave radiation")}
            # Atmospheric condition
            air_temp = self._parse_format(self.heat_exchange["air_temperature"], "Air temperature")
            rel_humidity = self._parse_format(self.heat_exchange["relative_humidity"], "Relative humidity")
            clear_coef = self._parse_format(self.heat_exchange["clearness_coefficient"], "Clearness coefficient")
            if self.heat_exchange["type_of_short_wave_radiation"] == 1 or self.heat_exchange["type_of_long_wave_radiation"] == 1:
                atmospheric = air_temp | rel_humidity | clear_coef
            else:
                atmospheric = air_temp | rel_humidity
            # Ground heat
            if self.heat_exchange["type_of_ground_heat"] == 0:
                ground = "Not included"
            elif self.heat_exchange["type_of_ground_heat"] == 1:
                ground = {"Distance below ground": str(self.heat_exchange["distance_below_ground"]) + " m"}
                thermal = self._parse_format(self.heat_exchange["THERMAL_CONDUCTIVITY"], "Thermal conductivity")
                ground_temp = self._parse_format(self.heat_exchange["GROUND_TEMPERATURE"], "Ground temperature")
                ground = ground | thermal | ground_temp
            output = {"Latent heat": latent, "Sensible heat": sensible, "Short wave radiation": short_wave, "Long wave radiation": long_wave, "Atmospheric condition": atmospheric, "Ground heat": ground}
            
        return output

    def __parse_precipitation_evaporation(self):
        if self.precipitation_evaporation["type_of_precipitation"] == 1:
            precipitation = {"Precipitation type": "Ambient water temperature"}
        elif self.precipitation_evaporation["type_of_precipitation"] == 2:
            precipitation = {"Precipitation type": "Specified temperature"} | self._parse_format(self.precipitation_evaporation["PRECIPITATION"], "Temperature", has_soft_time=True)
        if self.precipitation_evaporation["type_of_evaporation"] == 1:
            evaporation = {"Evaporation type": "Ambient water temperature"}
        elif self.precipitation_evaporation["type_of_evaporation"] == 2:
            evaporation = {"Evaporation type": "Specified temperature"} | self._parse_format(self.precipitation_evaporation["EVAPORATION"], "Temperature", has_soft_time=True)
        return {"Precipitation": precipitation, "Evaporation": evaporation}

    def __parse_infiltration(self):
        if self.infiltration["type_of_infiltration_temperature"] == 1:
            infiltration = {"Infiltration type": "Ambient water temperature"}
        elif self.infiltration["type_of_infiltration_temperature"] == 2:
            infiltration = {"Infiltration type": "Specified temperature"} | self._parse_format(self.infiltration["INFILTRATION"], "Temperature", has_soft_time=True)
        return infiltration

    def __parse_initial_conditions(self):
        return {"Temperature": self._parse_format(self.initial_conditions["TEMPERATURE"], "Temperature")}

class PFS_Turbulence(utils):
    def __init__(self, turbulence, boundaries):
        super().__init__()
        self.turbulence = turbulence
        self.boundaries = boundaries
        self.equation = self._parse_equation()
        self.model = self.equation["Model"]
        self.dispersion = self._parse_dispersion()
        self.initial_condition = self._parse_initial_conditions()
        self.attributes = {"Governing equation": self.equation,
                           **PFS_Solution_Technique(self.turbulence["SOLUTION_TECHNIQUE"], module="salinity_temperature").attributes,
                           "Dispersion": self.dispersion,
                           "Initial Conditions": self.initial_condition,
        }
        boundary_keys = {"k-epsilon": ["KINETIC_ENERGY", "DISSIPATION_OF_KINETIC_ENERGY"], "k-omega": ["KINETIC_ENERGY", "SPECIFIC_DISSIPATION_RATE"]}
        variables = {"KINETIC_ENERGY": "Turbulent kinetic energy", "DISSIPATION_OF_KINETIC_ENERGY": "Dissipation of turbulent kinetic energy", "SPECIFIC_DISSIPATION_RATE": "Specific dissipation rate"}
        self._parse_boundaries(boundary_keys[self.model], variables)
    
    def _parse_equation(self):
        equation = self.turbulence["EQUATION"]
        if equation["model"] == 1:
            output = {"Model": "k-epsilon",
                      "Empirical constants": {"c1e": equation["c1e"],
                                              "c2e": equation["c2e"],
                                              "c3e": equation["c3e"],
                                              "Prandtl number": equation["prandtl_number"],
                                              "cmy": equation["cmy"]},
                        "Minimum turbulent kinetic energy": str(equation["minimum_kinetic_energy"]) + " m²/s²",
                        "Maximum turbulent kinetic energy": str(equation["maximum_kinetic_energy"]) + " m²/s²",
                        "Minimum dissipation of turbulent kinetic energy": str(equation["minimum_dissipation_of_kinetic_energy"]) + " m²/s³",
                        "Maximum dissipation of turbulent kinetic energy": str(equation["maximum_dissipation_of_kinetic_energy"]) + " m²/s³",
                        "Vegetation coefficient - epsilon": str(equation["vegetation_coefficient_epsilon"]),
                        "Damping": "Included" if equation["Ri_damping"] == 1 else "Not included"}
        elif equation["model"] == 2:
            output = {"Model": "k-omega",
                      "Empirical constants": {"alpha": equation["alpha"],
                                              "beta_0": equation["beta_0"],
                                              "sigma_do": equation["sigma_do"],
                                              "Prandtl number": equation["prandtl_number_KO"],
                                              "beta_k": equation["beta_k"],},
                        "Minimum turbulent kinetic energy": str(equation["minimum_kinetic_energy"]) + " m²/s²",
                        "Maximum turbulent kinetic energy": str(equation["maximum_kinetic_energy"]) + " m²/s²",
                        "Minimum specific dissipation rate": str(equation["minimum_specific_dissipation_rate"]) + " 1/s",
                        "Maximum specific dissipation rate": str(equation["maximum_specific_dissipation_rate"]) + " 1/s",
                        "Vegetation coefficient - omega": str(equation["vegetation_coefficient_omega"])}
        output["Vegetation coefficient - k"] = str(equation["vegetation_coefficient_k"])
        return output

    def _parse_dispersion(self):
        dispersion = self.turbulence["DIFFUSION"]
        if self.model == "k-epsilon":
            return {"Turbulent kinetic energy": {"Horizontal sigma": dispersion["sigma_k_h"],
                                                "Vertical sigma": dispersion["sigma_k"]},
                    "Dissipation of turbulent kinetic energy": {"Horizontal sigma": dispersion["sigma_e_h"],
                                                                "Vertical sigma": dispersion["sigma_e"]}}
        elif self.model == "k-omega":
            return {"Turbulent kinetic energy": {"Horizontal sigma": dispersion["sigma_turbulent_kinetic_energy_h"],
                                                "Vertical sigma": dispersion["sigma_turbulent_kinetic_energy_v"]},
                    "Specific dissipation rate": {"Horizontal sigma": dispersion["sigma_specific_dissipation_rate_h"],
                                                "Vertical sigma": dispersion["sigma_specific_dissipation_rate_v"]}}

    def _parse_initial_conditions(self):
        initial = self.turbulence["INITIAL_CONDITIONS"]
        if self.model == "k-epsilon":
            return {"Turbulent kinetic energy": self._parse_format(initial["KINETIC_ENERGY"], "Turbulent kinetic energy"),
                    "Dissipation of turbulent kinetic energy": self._parse_format(initial["DISSIPATION_OF_KINETIC_ENERGY"], "Dissipation of turbulent kinetic energy")}
        elif self.model == "k-omega":
            return {"Turbulent kinetic energy": self._parse_format(initial["KINETIC_ENERGY"], "Turbulent kinetic energy"),
                    "Specific dissipation rate": self._parse_format(initial["SPECIFIC_DISSIPATION_RATE"], "Specific dissipation rate")}

    def _parse_boundaries(self, keys, variables):
        for key in self.boundaries.keys():
            
            for model_key in keys:
                data = self.turbulence["BOUNDARY_CONDITIONS"][key][model_key]
                if len(data.keys()) == 0:
                    continue
                if "Turbulence Information" not in self.boundaries[key].keys():
                    self.boundaries[key]["Turbulence Information"] = {}
                if data["type"] == 1:
                    output = {"Type": "Land"}
                elif data["type"] == 2:
                    output = {"Type": "Specified value",
                                **self._parse_format(data, variable=variables[model_key]),
                                "Type of soft start": "Linear variation" if data["type_of_soft_start"] == 1 else "Sine variation"}
                    if data["format"] == 1:
                        output["Type of interpolation in time"] = "Linear" if data["type_of_time_interpolation"] == 1 else "Piecewise cubic"
                    if data["format"] == 3:
                        output["Type of interpolation in time"] = "Linear" if data["type_of_time_interpolation"] == 1 else "Piecewise cubic"
                        output["Type of interpolation in space"] = "Normal" if data["type_of_space_interpolation"] == 1 else "Reverse order"
                elif data["type"] == 3:
                    output = {"Type": "Zero gradient"}
                self.boundaries[key]["Turbulence Information"][variables[model_key]] = output

class PFS_Hydrodynamic(utils):
    def __init__(self, hydrodynamic, boundaries):
        super().__init__()
        self.hydrodynamic = hydrodynamic
        self.boundaries = boundaries
        if hydrodynamic["mode"] != 2:
            return {"Hydrodynamic module": "Not included"}
        self.governing_equation = PFS_Equation(hydrodynamic["EQUATION"]).attributes
        self.solution_technique = PFS_Solution_Technique(hydrodynamic["SOLUTION_TECHNIQUE"], module="hydrodynamic").attributes
        self.depth_correction = {"Depth correction": "Not included"} if hydrodynamic["DEPTH"]["type"] == 0 else {"Depth correction type": "Specified bed level change", **self._parse_format(hydrodynamic["DEPTH"], variable="Depth correction", key="constant_value")}
        self.flood_dry = self._parse_Flood_Dry(hydrodynamic["FLOOD_AND_DRY"])
        self.density = self._parse_density(hydrodynamic["DENSITY"])
        self.eddy = PFS_Eddy(hydrodynamic["EDDY_VISCOSITY"]).attributes
        self.bed_resistence = Bed_Resistence(hydrodynamic["BED_RESISTANCE"]).attributes
        self.vegetation = "Not included" if self.hydrodynamic["VEGETATION"]["type"] == 0 else "Included"
        self.coriolis = self._parse_Coriolis(hydrodynamic["CORIOLIS"])
        self.wind = Wind_Forcing(hydrodynamic["WIND_FORCING"]).attributes
        self.ice = self._parse_Ice_Coverage(hydrodynamic["ICE"])
        self.tide = self._parse_Tidal_Potential(hydrodynamic["TIDAL_POTENTIAL"])
        self.precipitation_evaporation = self._parse_precipitation_evaporation(hydrodynamic["PRECIPITATION_EVAPORATION"])
        self.infilitration = self._parse_infiltration(hydrodynamic["INFILTRATION"])
        self.wave_radiation = {"Wave radiation": "Not included"} if hydrodynamic["RADIATION_STRESS"]["type"] == 0 else {"Wave radiation": self._parse_format(hydrodynamic["RADIATION_STRESS"], variable="Wave radiation", key="constant_value", has_soft_time=True)}
        self.sources = PFS_Source(hydrodynamic["SOURCES"]).attributes
        self.n_sources = self.sources["Number of sources"]
        self.initial_conditions = self._parse_Initial_Condition(hydrodynamic["INITIAL_CONDITIONS"])
        self.boundaries = PFS_Boundaries(hydrodynamic["BOUNDARY_CONDITIONS"], self.boundaries).boundaries
        tmp = PFS_Salinity_Temperature(hydrodynamic["TEMPERATURE_SALINITY_MODULE"], self.boundaries, self.sources)
        self.salinity_temperature = tmp.attrubutes
        self.boundaries = tmp.boundaries
        tmp = PFS_Turbulence(hydrodynamic["TURBULENCE_MODULE"], self.boundaries)
        self.turbulence = tmp.attributes
        self.boundaries = tmp.boundaries
        self.attributes = {"Governing equation": self.governing_equation,
                           "Solution technique": self.solution_technique,
                           "Depth correction": self.depth_correction,
                           "Flood and dry": self.flood_dry,
                           "Density": self.density,
                           "Eddy viscosity": self.eddy,
                           "Bed resistence": self.bed_resistence,
                           "Vegetation": self.vegetation,
                           "Coriolis": self.coriolis,
                           "Wind forcing": self.wind,
                           "Ice coverage": self.ice,
                           "Tidal potential": self.tide,
                           "Precipitation and evaporation": self.precipitation_evaporation,
                           "Infiltration": self.infilitration,
                           "Wave radiation": self.wave_radiation,
                           "Initial conditions": self.initial_conditions,
                           "Salinity and temperature": self.salinity_temperature,
                           "Turbulence": self.turbulence}
                           


    def _parse_Flood_Dry(self, flood_dry):
        if flood_dry["type"] == 0:
            return "Not included"
        else:
            return {"Flooding depth": flood_dry["mass_depth"], "Drying depth": flood_dry["drying_depth"]}
        
    def _parse_density(self, density):
        types = {0: "Barotropic", 1: "Function of temperature and salinity", 2: "Function of temperature", 3: "Function of salinity"}
        output = {"Density type": types[density["type"]]}
        reference_data = {
            1: {"Reference Temperature": density["temperature_reference"], "Reference Salinity": density["salinity_reference"]},
            2: {"Reference Temperature": density["temperature_reference"]},
            3: {"Reference Salinity": density["salinity_reference"]}
        }
        output.update(reference_data.get(density["type"], {}))
        return output
    
    def _parse_Coriolis(self, coriolis):
        if coriolis["type"] == 0:
            return "No coriolis force"
        elif coriolis["type"] == 1:
            return {"Coriolis forcing format": "Constant coriolis forcing in space", "Reference latitude": coriolis["latitude"]}
        elif coriolis["type"] == 2:
            return {"Coriolis forcing format": "Varying coriolis forcing in space"}
        
    def _parse_Ice_Coverage(self, ice):
        if ice["type"] == 0:
            return {"Ice coverage": "Not included"}
        else:
            if ice["type"] == 1:
                output = {"Ice coverage type": "Specified ice concentration", "Critical ice concentration": ice["c_cut_off"]}
            elif ice["type"] == 2:
                output = {"Ice coverage type": "Specified ice thickness"}
            elif ice["type"] == 3:
                output = {"Ice coverage type": "Specified ice concentration and thickness", "Critical ice concentration": ice["c_cut_off"]}
            ice_roughness = ice["ROUGHNESS"]
            if ice_roughness["type"] == 0:
                output.update({"Ice roughness": "Not included"})
            else:
                output.update(self._parse_format(ice_roughness, variable="Ice roughness height", key="constant_value"))
            return output
        
    def _parse_Tidal_Potential(self, tide):
        if tide["type"] == 0:
            return "Not included"
        else:
            if tide["format"] == 0:
                n_constituents = tide["number_of_constituents"]
                constituents = " ".join([tide[f"CONSTITUENT_{i+1}"]["name"] for i in range(n_constituents)])
                return {"Number of constituents": n_constituents, "Constituents": constituents}
            elif tide["format"] == 1:
                return "Included"

    def _parse_precipitation_evaporation(self, precipitation_evaporation):
        precipitation = self._parse_format(precipitation_evaporation["PRECIPITATION"], variable="Precipitation", key="constant_value", has_soft_time=True)
        evaporation = self._parse_format(precipitation_evaporation["EVAPORATION"], variable="Evaporation", key="constant_value", has_soft_time=True)
        if precipitation_evaporation["type_of_precipitation"] == 0:
            prep = {"Precipitation": "Not included"}
        elif precipitation_evaporation["type_of_precipitation"] == 1:
            prep = {"Precipitation type": "Specified precipitation", **precipitation}
        elif precipitation_evaporation["type_of_precipitation"] == 2:
            return {"Precipitation type": "Net precipitation", **precipitation}
        if precipitation_evaporation["type_of_evaporation"] == 0:
            evap = {"Evaporation": "Not included"}
            return prep | evap
        elif precipitation_evaporation["type_of_evaporation"] == 1:
            evap = {"Evaporation type": "Specified evaporation", **evaporation}
            return prep | evap
        elif precipitation_evaporation["type_of_evaporation"] == 2:
            evap = {"Evaporation type": "Computed evaporation"}
            return prep | evap

    def _parse_infiltration(self, infiltration):
        if infiltration["type"] == 0:
            return {"Infiltration": "Not included"}
        elif infiltration["type"] == 1:
            return {"Infiltration type": "Net infiltration rate", **self._parse_format(infiltration, variable="Net infiltration rate", key="constant_value")}
        elif infiltration["type"] == 2:
            return {"Infiltration type": "Constant infiltration with capacity", **self._parse_format(infiltration, variable="Infiltration capacity", key="constant_value")}

    def _parse_Initial_Condition(self, initial_condition):
        if initial_condition["type"] == 1:
            return {"Initial condition type": "Constant initial condition",
                    "Initial surface elevation": str(initial_condition["surface_elevation_constant"]) + " m",
                    "Initial u-velocity": str(initial_condition["u_velocity_constant"]) + " m/s",
                    "Initial v-velocity": str(initial_condition["v_velocity_constant"]) + " m/s",
                    "Initial w-velocity": str(initial_condition["w_velocity_constant"]) + " m/s"}
        elif initial_condition["type"] == 2:
            return {"Initial condition type": "Varying surface elevation in space"}
        elif initial_condition["type"] == 3:
            return {"Initial condition type": "Varying water depth and velocities in space"}

class PFS_ST_Bed_Resistence(utils):
    def __init__(self, bed):
        super().__init__()
        self.bed = bed
        beds = {0: "No bed resistance", 1: "Chezy number", 2: "Manning number", 3: "Alluvial resistance", 4: "Resistance from HD simulations"}
        keys = {0: None, 1: "CHEZY_NUMBER", 2: "MANNING_NUMBER", 3: "ALLUVIAL_RESISTANCE", 4: None}
        self.type = beds[bed["type"]]
        self.attributes = {"Bed resistance type": self.type}
        if bed["type"] in [1, 2, 3]:
            self.attributes.update(self._parse_format(bed[keys[bed["type"]]], variable=beds[bed["type"]], key="constant_value"))
        if bed["type"] == 3:
            self.attributes.update(self._parse_alluvial_resistance(bed[keys[bed["type"]]]))
        
    def _parse_alluvial_resistance(self, alluviaul):
        return {"Alluvial resistance power": alluviaul["resistance_power"],
                "Minimum alluvial resistance": alluviaul["minimum_resistance"],
                "Maximum alluvial resistance": alluviaul["maximum_resistance"],
                }

class PFS_ST_Morphology(utils):
    def __init__(self, morphology, boundaries):
        super().__init__()
        self.morphology = morphology
        self.boundaries = boundaries
        self.boundary_conditions = self.morphology["BOUNDARY_CONDITIONS"]
        self.model_definition = self._parse_model_definition()
        self.bank_erosion = self._parse_slope_failure()
        self._parse_boundaries()
        self._parse_bed_level_sources()
        self.attributes = {"Model definition": self.model_definition,
                           "Bank erosion": self.bank_erosion,
                           "Bed level sources": self.bed_level_sources}
        
    def _parse_boundaries(self):
        for key in self.boundaries.keys():
            boundary = self.boundary_conditions[key]
            if len(boundary.keys()) == 0:
                continue
            types = {1: "Land", 2: "Zero sediment flux gradient", 3: "Zero sediment flux gradient for outflow, zero bed level change for inflow"}
            self.boundaries[key]["Morphology Information"] = {"Type": types[boundary["type"]]}

    def _parse_bed_level_sources(self):
        sources = self.morphology["BED_LEVEL_SOURCES"]
        self.bed_level_sources = {}
        if sources["number_of_sources"] == 0:
            return "No bed level sources"
        for i in range(sources["number_of_sources"]):
            source = sources[f"BED_LEVEL_SOURCE_{i+1}"]
            types = {1: "Standard", 2: "Dredging"}
            variables = {1: "Added rate of bed level change", 2: "Maximum allowed bed level"}
            output = {"Name": source["name"],
                      "Type": types[source["type"]],
                      **self._parse_format(source, variable=variables[source["type"]], key="constant_value")}
            if source["type"] == 2 and source["DISTRIBUTION_FACTOR"]["type"] == 2:
                output.update(self._parse_format(source["DISTRIBUTION_FACTOR"], variable="Distribution factor", key="constant_value"))
            self.bed_level_sources[f"Source {i+1}"] = output
             
    def _parse_model_definition(self):
        model = self.morphology["MODEL_DEFINITION"]
        return {"Maximum bed level change": str(model["max_bed_level_change"]) + " m/d",
                  "Feedback on hydrodynamic, waves, and sand transport": "Included" if model["include_morphology_update"] == 1 else "Not included",
                  **self._parse_format(model["SPEEDUP_FACTOR"], variable="Speedup factor", key="constant_value", has_soft_time=True)}
        
    def _parse_slope_failure(self):
        slope = self.morphology["BANK_EROSION"]
        types = {0: "Not included", 1: "Simple bank erosion", 2: "Extended bank erosion", 3: "General slope failure", 4: "General slope failure including dry elements"}
        if slope["type"] == 0:
            return {"Bank erosion": "Not included"}
        else:
            return {"Bank erosion": types[slope["type"]],
                    **self._parse_format(slope["ANGLE_OF_REPOSE"], variable="Angle of repose", key="constant_value")}

class PFS_Sand_Transport(utils):
    def __init__(self, ST, boundaries, sources):
        super().__init__()
        self.ST = ST
        self.boundaries = boundaries
        self.sources = sources
        self.n_layers = 0
        self.sources_condition = self.ST["SOURCES"]
        self.varying_layer_thickness = False
        self.equilibrium = False
        self.model_definition = self._parse_model_definition()
        self.solution_technique = PFS_Solution_Technique(ST["SOLUTION_TECHNIQUE"], module="salinity_temperature").attributes
        self.sediment_properties = self._parse_sediment_properties()
        self.bed_resistence = PFS_ST_Bed_Resistence(ST["BED_RESISTANCE"]).attributes
        self.forcing = self._parse_forcing()
        self._parse_sources()
        self.initial_conditions = self._parse_initial_conditions()
        self.boundary_conditions = self.ST["BOUNDARY_CONDITIONS"]
        self._parse_boundaries()
        tmp = PFS_ST_Morphology(ST["MORPHOLOGY"], self.boundaries)
        self.morphology = tmp.attributes
        self.boundaries = tmp.boundaries

        self.attributes = {"Model definition": self.model_definition,
                           "Solution technique": self.solution_technique,
                           "Sediment properties": self.sediment_properties,
                           "Bed resistence": self.bed_resistence,
                           "Forcing": self.forcing,
                           "Initial conditions": self.initial_conditions,
                           "Morphology": self.morphology}

    def _parse_sources(self):
        for i in range(self.sources["Number of sources"]):
            source = self.sources_condition[f"SOURCE_{i+1}"]
            types = {1: "Specified concentration", 2: "Excess concentration"}
            self.sources[f"Source {i+1}"]["Sand Transport Information"] = {"Type": types[source["SSC_FRACTION_1"]["type"]],
                                                                    **self._parse_format(source["SSC_FRACTION_1"], variable="Concentration")}

    def _parse_boundaries(self):
        for key in self.boundaries.keys():
            boundary = self.boundary_conditions[key]
            if len(boundary.keys()) == 0:
                continue
            boundary = boundary["SSC_FRACTION_1"]
            if boundary["type"] == 1:
               output = {"Type": "Land"}
            elif boundary["type"] == 2:
                output = {"Type": "Specified concentration",
                            **self._parse_format(boundary["SSC_FRACTION_1"], variable="Concentration"),
                            "Type of soft start": "Linear variation" if boundary["type_of_soft_start"] == 1 else "Sine variation"}
                if boundary["format"] == 1:
                    output["Type of interpolation in time"] = "Linear" if boundary["type_of_time_interpolation"] == 1 else "Piecewise cubic"
                if boundary["format"] == 3:
                    output["Type of interpolation in time"] = "Linear" if boundary["type_of_time_interpolation"] == 1 else "Piecewise cubic"
                    output["Type of interpolation in space"] = "Normal" if boundary["type_of_space_interpolation"] == 1 else "Reverse order"
            elif boundary["type"] == 3:
                output = {"Type": "Zero gradient"}
            elif boundary["type"] == 4:
                output = {"Type": "Equilibrium conditions"}
            self.boundaries[key]["Sand Transport Information"] = output

    def _parse_model_definition(self):
        model = self.ST["MODEL_DEFINITION"]
        model_type = model["model"]
        multi_fraction_multi_layer = model["multi_fraction_multi_layer"]
        transport_description = "Equilibrium" if model["transport_description"] == 0 else "Non-equilibrium"
        helical_flow_description = model["helical_flow_description"]
        formulation_forcing = "Depth-averaged velocity" if model["formulation_forcing"] == 1 else "Bed shear stress"
        number_of_fractions = model["number_of_fractions"]
        number_of_fractions_multi = model["number_of_fractions_multi"]
        equilibrium_layer_thickness = model["equilibrium_layer_thickness"]
        number_of_layers = f"Varuting layer thickness with threshold thickness of {equilibrium_layer_thickness} m" if model["number_of_layers"] > 0 else "Constant layer thickness"
        number_of_layers_multi = model["number_of_layers_multi"]
        total_load_factor = model["total_load_factor"]
        maximum_thickness_surface_layer = model["maximum_thickness_surface_layer"]
        maximum_thickness_sublayer = model["maximum_thickness_sublayer"]
        data_file = model["data_file"]
        fraction = model["FRACTION_1"]
        formulas = {0: "No bed", 1: "Engelund-Hansen", 2: "van Rijn", 3: "Engelund-Fredsoe", 4: "Meyer-Peter-Muller", 5: "Kovacs and Parker", 6: "Engelund-Fredsoe (Extended)"}
        bed_load_formula = formulas[fraction["bed_load_formula"]]
        bed_load_factor = fraction["bed_load_factor"]
        bed_concentration_formula = fraction["bed_concentration_formula"]
        suspended_load_formula = formulas[fraction["suspended_load_formula"]]
        suspended_load_factor = fraction["suspended_load_factor"]
        maximum_value = str(fraction["maximum_value"]) + " kg/m³"
        type_of_shear_stress = fraction["type_of_shear_stress"]
        static_friction_coefficient = fraction["static_friction_coefficient"]
        dynamic_friction_coefficient = fraction["dynamic_friction_coefficient"]
        friction_velocity_factor = fraction["friction_velocity_factor"]
        if model_type == 1:
            if multi_fraction_multi_layer == 0:
                self.model_type = 0
                self.varying_layer_thickness = True if model["number_of_layers"] > 0 else False
                self.equilibrium = True if model["transport_description"] == 0 else False
                output = {"Model type": "Pure current",
                          "Layer thickness": number_of_layers,
                          "Forcing parameter": formulation_forcing,
                          "Transport description": transport_description,
                          "Bed load formula": bed_load_formula,
                          "Bed load factor": bed_load_factor,
                          "Suspension load formula": suspended_load_formula,
                          "Suspension load factor": suspended_load_factor,
                          "Maximum concentration": maximum_value}
                
                if bed_load_formula in ["Kovacs and Parker", "Engelund-Fredsoe (Extended)"]:
                    output["Static friction coefficient"] = static_friction_coefficient
                    output["Dynamic friction coefficient"] = dynamic_friction_coefficient
                    output["Friction velocity factor"] = friction_velocity_factor
            elif multi_fraction_multi_layer == 1:
                self.model_type = 1
                self.n_layers = number_of_layers_multi
                output = {"Model type": "Pure current - Multi-fraction/multi-layer",
                          "Number of fractions": number_of_fractions_multi,
                          "Number of layers": number_of_layers_multi,
                          "Thickness surface layer": str(maximum_thickness_surface_layer) + " m",
                          "Thickness sublayer": str(maximum_thickness_sublayer) + " m",
                          "Forcing parameter": formulation_forcing}
        elif model_type == 2:
            self.model_type = 2
            self.varying_layer_thickness = True if model["number_of_layers"] > 0 else False
            output = {"Model type": "Wave and current",
                      "Layer thickness": number_of_layers,
                      "Forcing parameter": formulation_forcing}
        return output

    def _parse_sediment_properties(self):
        properties = self.ST["SEDIMENT_PROPERTIES"]
        porosity = properties["porosity"]
        data = properties["SEDIMENT_DATA"]
        grain_diameter = self._parse_format(data, variable="Grain diameter", key="grain_diameter")
        grading_coefficient = data["grading_coefficient"]
        relative_density = data["relative_density"]
        critical_Shields_parameter = data["critical_Shields_parameter"]
        grain_diameter_for_fractions = data["grain_diameter_for_fractions"]
        density_for_fractions = data["density_for_fractions"]
        if self.model_type == 0:
            return {"Porosity": porosity,
                    **grain_diameter,
                    "Relative density": relative_density,
                    "Critical Shields parameter": critical_Shields_parameter}
        elif self.model_type == 1:
            return {"Porosity": porosity,
                    "Grain diameter for fractions": [str(elem) + " mm" for elem in grain_diameter_for_fractions],
                    "Density for fractions": [str(elem) + " kg/m³" for elem in density_for_fractions]}
        elif self.model_type == 2:
            if data["format"] == 0:
                return {"Porosity": porosity,
                        **grain_diameter,
                        "Grading coefficient": grading_coefficient}
            else:
                return {"Porosity": porosity,
                        **grain_diameter}

    def _parse_forcing(self):
        if self.model_type in [0, 1]:
            return "Not included"
        forcing = self.ST["FORCINGS"]["WAVES"]["WAVE_FIELD"]
        output = {"Forcing": "Wave field",
                  "Wave height field": "Signigicant wave height" if forcing["type_of_wave_height"] == 2 else "RMS wave height",
                  "Wave period field": "Peak wave period" if forcing["type_of_wave_period"] == 1 else "Mean wave period"}
        if forcing["format"] == 0:
            output.update({"Wave field format": "Constant wave field",
                           "Wave height": str(forcing["wave_height"]) + " m",
                           "Wave period": str(forcing["wave_period"]) + " s",
                           "Wave direction": str(forcing["wave_direction"]) + " degrees"})
        elif forcing["format"] == 1:
            output.update({"Wave field format": "Varying in time, constant in space"})
        elif forcing["format"] == 2:
            output.update({"Wave field format": "Constant in time, varying in space"})
        elif forcing["format"] == 3:
            output.update({"Wave field format": "Varying in time and space"})
        return output

    def _parse_dispersion(self):
        dispersion = self.ST["DISPERSION"]["HORIZONTAL_DISPERSION"]["SSC_FRACTION_1"]
        if dispersion["type"] == 0:
            return "Not included"
        elif dispersion["type"] == 1:
            return {"Formulation": "Dispersion coefficient",
                    **self._parse_format(dispersion["DISPERSION_COEFFICIENT"], variable="Dispersion coefficient", key="constant_value")}
        elif dispersion["type"] == 2:
            return {"Formulation": "Scaled eddy viscosity",
                    **self._parse_format(dispersion["SCALED_EDDY_VISCOSITY"], variable="Scaled eddy viscosity", key="sigma")}

    def _parse_initial_conditions(self):
        initial = self.ST["INITIAL_CONDITIONS"]
        pure_current = self.__parse_pure_current_initial(initial["SSC_FRACTION_1"])
        multi_layer = self.__parse_multi_layer_initial(initial)
        if self.model_type == 0:
            return pure_current
        elif self.model_type == 1:
            return multi_layer

    def __parse_pure_current_initial(self, pure_current):
        if pure_current["type"] == 2:
            return "Equilibrium conditions" 
        elif pure_current["type"] == 1:
            return {"Type": "Specified concentration",
                    **self._parse_format(pure_current, variable="Concentration", key="constant_value")}

    def __parse_multi_layer_initial(self, initial):
        layers_initial_condition = {}
        for i in range(self.n_layers):
            layer = initial[f"BED_THICKNESS_MULTI_LAYER_{i+1}"]
            output =  {**self._parse_format(layer, variable="Thickness", key="constant_value"),
                       **self._parse_format(layer, variable="Composition", key="constant_value")}
            layers_initial_condition[f"Layer {i+1}"] = output
        return layers_initial_condition
            
class PFS_Parser:
    def __init__(self, fname):
        self.fname = fname
        self.pfs = mikeio.read_pfs(fname)["FemEngineHD"]
        self.modules = [key for key in self.pfs["MODULE_SELECTION"].keys() if self.pfs["MODULE_SELECTION"][key] == 2]
        self.domain = PFS_Domain(self.pfs["DOMAIN"])
        self.boundaries = self.domain.boundaries
        self.time = PFS_Time(self.pfs["TIME"])
        self.hydrodynamic = PFS_Hydrodynamic(self.pfs["HYDRODYNAMIC_MODULE"], self.boundaries)
        self.sources = self.hydrodynamic.sources
        self.boundaries = self.hydrodynamic.boundaries
        if "mode_of_sand_transport_module" in self.modules:    
            self.sand_transport = PFS_Sand_Transport(self.pfs["SAND_TRANSPORT_MODULE"], self.boundaries, self.sources)
            self.sources = self.sand_transport.sources
            self.boundaries = self.sand_transport.boundaries
        
    def parse(self):
        output = {"Domain": self.domain.attributes,
                "Time": self.time.attributes,
                "Hydrodynamics": self.hydrodynamic.attributes}
        if "mode_of_sand_transport_module" in self.modules:
            output["Sand Transport"] = self.sand_transport.attributes
        output["Boundaries"] = self.boundaries
        output["Sources"] = self.sources
        return output
    
    def to_excel(self, fname):
        data = self.parse()
        self.row = 1
        self.heading = 1
        self.indent = ""
        self.fonts = {1: Font(bold=True, size=16),
                      2: Font(bold=True, size=15),
                      3: Font(bold=True, size=14),
                      4: Font(bold=True, size=13),
                      5: Font(bold=True, size=12)}
        self.wb = Workbook()
        self.ws = self.wb.active
        self._write_dict(data)
        self.wb.save(fname)
        self.wb.close()

    
    def _write_dict(self, dictionary):
        for key in dictionary.keys():
            if isinstance(dictionary[key], dict):
                self.ws.cell(row=self.row, column=1).value = self.indent+str(key)
                self.ws.cell(row=self.row, column=1).font = self.fonts[self.heading]
                self.heading += 1
                self.row += 1
                self.indent += "    "
                self._write_dict(dictionary[key])
            else:
                self.ws.cell(row=self.row, column=1).value = self.indent+str(key)
                self.ws.cell(row=self.row, column=2).value = str(dictionary[key])
                print(key, dictionary[key], self.row)
                self.row += 1
        self.heading -= 1
        self.indent = self.indent[:-4]





sample_files = [f for f in os.listdir("sample_input_files")]
fname = os.path.join("sample_input_files", sample_files[9])
print(fname)
pfs = PFS_Parser(fname)
print(pfs.to_excel("Test.xlsx"))